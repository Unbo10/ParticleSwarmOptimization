import os

import numpy as np
import openpyxl as px
import openpyxl.styles as px_styles
import openpyxl.worksheet.worksheet as px_worksheet
import pandas as pd
from psopackage.graphics.colors import Color

class Data:
    """
    This class is used to store the optimization results in an Excel file. It is used to store the history of the particles and the gbest indexes of each optimization.
    
    """
    def __init__(self, excel_file_name: str) -> None:
        """
        Initializes the Data object with the given Excel file name.
        
        Args:
            excel_file_name (str): The name of the Excel file (without extension and directory).
        """    

        # List to store the history of particles for each optimization
        self.__particle_history: list[pd.DataFrame] = []
        
        # List to store the gbest indexes for each optimization
        self.__gbest_history: list[list[int]] = []
        
        # Counter to keep track of the number of optimizations
        self.__number_of_optimizations: int = 0
        
        # Name of the Excel file (without extension)
        self.__xlsx_name: str = excel_file_name
        
        # Full path to the Excel file
        self.__xlsx_path: str = f"database/optimization_results/{self.__xlsx_name}.xlsx"
        
        # * The two dots are needed if the GUI is directly executed.
        # ! For now, the execution will continue to be done in the gui.py file, but the final version MUST CHANGE the paths to execute everything from the main.py file.
        # * Already done
        
    def append_gbest_indexes(self, optimization_gbest_indexes: list[int]) -> None:
        """
        The gbest indexes are appended to the gbest history list. This method is called after each optimization.
        
        """
        self.__gbest_history.append(optimization_gbest_indexes)
        self.__number_of_optimizations += 1
    
    def append_optimization(self, optimization_df: pd.DataFrame) -> None:
        """
        This method styles the DataFrame and appends it to the Excel file. Creates the Excel file if it does not exist.
        
        """
        # TODO: Test the whole class with multiple sessions and files. Also, update and add documentation and the class diagram.
        # * Seems to be working alright
        if "optimization_results" not in os.listdir("database"):
            os.mkdir("database/optimization_results")
            
        # ? We may want to implement this using with
        
        # * Creating the Excel file if it does not exist
        if not os.path.exists(self.__xlsx_path):
            self.create_spreadsheet()
        
        # * Open the workbook and CREATE THE SHEET
        workbook: px.Workbook = px.load_workbook(self.__xlsx_path)
        sheet_number: int = self.__number_of_optimizations
        sheet_name: str = f"Optimization {sheet_number}"
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(title=f"Optimization {sheet_number}")
        else:
            pass

        sheet: px_worksheet.Worksheet = workbook[sheet_name]
        # * Freezing the first two written rows
        # ? Should it only be the second one?
        sheet.freeze_panes = "A3"
        sheet.freeze_panes = "A4"

        # * Variables for styling the BORDERS
        title_side = px_styles.Side(style="slantDashDot", color=Color.siders) 
        # ! Replace for a Color attribute once the branch is merged
        # * Done
        title_border = px_styles.Border(left=title_side, right=title_side,
            top=title_side, bottom=title_side)
        header_side = px_styles.Side(style="thick", color=Color.siders) 
        # ! Replace for a Color attribute once the branch is merged
        # * Done
        header_border = px_styles.Border(left=header_side,
            right=header_side, top=header_side, bottom=header_side)
        common_side = px_styles.Side(style="thin", color=Color.common_side) 
        # ! Replace for a Color attribute once the branch is merged
        # * Done
        common_border = px_styles.Border(left=common_side,
            right=common_side, top=common_side, bottom=common_side)
        spacing_side = px_styles.Side(style="thin", color=Color.siders) 
        # ! Replace for a Color attribute once the branch is merged
        # * Done
        spacing_border = px_styles.Border(left=spacing_side,
            right=spacing_side, top=common_side, bottom=common_side)
        
        # * Setting the same FONT and size for all the cells
        # * It is only overwrittern in the headers (because of the font color)
        for row in sheet.iter_rows(min_row=3, max_row=optimization_df.shape[0] + 3, min_col=2, max_col=7):
            sheet.row_dimensions[row[0].row].height = 18
            for cell in row:
                cell.font = px_styles.Font(name="FreeMono", size=11,
                    bold=True)
                        
        # * Setting and styling the TITLE
        sheet.merge_cells("B2:G2")
        sheet.row_dimensions[2].height = 45
        sheet["B2"] = f"PSO optimization {self.__number_of_optimizations} results"
        title_cell = sheet.cell(row=2, column=2)
        title_cell.alignment = px_styles.Alignment(horizontal="center",
            vertical="center")
        title_cell.font = px_styles.Font(name="FreeMono", size=18,
            bold=True, color=Color.title_cell) 
        # ! Replace for a Color attribute once the branch is merged
        # * Done
        title_cell.fill = px_styles.GradientFill(stop=(Color.grad_fill_start, Color.grad_fill_stop),
            type="linear", degree=90) 
        # ! Replace for a Color attribute once the branch is merged. AND CHECK THE GRADIENT
        # * Done
        title_cell.border = title_border
        
        # * Naming the columns by naming and styling their HEADERS
        sheet["B3"] = "Iteration"
        sheet["C3"] = "Particle"
        sheet["D3"] = "Heuristic coordinates"
        sheet["E3"] = "Position coordinates"
        sheet["F3"] = "Velocity coordinates"
        sheet["G3"] = "Best personal position coordinates"
        for column in range(2, 8):
            cell = sheet.cell(row=3, column=column)
            cell.alignment = px_styles.Alignment(horizontal="center",
                vertical="center", wrap_text=True)
            cell.font = px_styles.Font(name="FreeMono", size=11, bold=True, color= Color.siders)
            cell.fill = px_styles.PatternFill(start_color= Color.cell_fill, end_color= Color.cell_fill, fill_type="solid") 
            # ! @U Replace for a Color attribute once the branch is merged
            # * Done
            cell.border = header_border
            sheet.row_dimensions[3].height = 40

        # * Converting the DataFrame to a tuple of tuples to access the data more easily
        optimization_records: tuple[tuple] = optimization_df.to_records(index=False)

        # * Setting the WIDTH of the columns according to the dimesion of the particles
        column_width: int = 6*(len(optimization_records[0][0]) + 1)
        for column in ["C", "D", "E", "F", "G"]:
            sheet.column_dimensions[column].width = column_width
        # number_of_iterations: int = optimization_df["Heuristic"].isna().sum() # ! Seems it is not needed
        sheet.column_dimensions["B"].width = 10

        # * Couting the AMOUNT OF PARTICLES
        number_of_particles: int = 0
        while optimization_records[number_of_particles][0] is not np.nan:
            number_of_particles += 1
        particle_index: int = 1
        # * Setting the particle indexes in the "C" column and styling them
        for row in range(optimization_df.shape[0]):
            cell = sheet.cell(row=row + 4, column=3)
            if optimization_records[row][0] is not np.nan:
                cell.value = particle_index
                cell.alignment = px_styles.Alignment(horizontal="center",
                    vertical="center")
                cell.fill = px_styles.PatternFill(start_color= Color.cell2_fill, 
                    end_color= Color.cell2_fill, fill_type="solid") 
                # ! Replace for a Color attribute once the branch is merged
                # * Done
                cell.border = common_border
                if particle_index == number_of_particles:
                    particle_index = 1
                else:
                    particle_index += 1
            else:
                # * Styling the spacing rows
                cell.border = spacing_border

        # * Merging the "B" column cells to show the ITERATION number
        starting_row: int = 4
        ending_row: str = number_of_particles + 3
        iteration_number: int = 0
        while starting_row < optimization_df.shape[0]:
            sheet.merge_cells(f"B{starting_row}:B{ending_row}")
            sheet[f"B{starting_row}"] = iteration_number
            sheet.cell(row=starting_row, column=2).fill = px_styles.PatternFill(start_color= Color.cell2_fill, end_color= Color.cell2_fill,
                fill_type="solid") 
            # ! Replace for a Color attribute once the branch is merged
            # * Done
            sheet.cell(row=starting_row, column=2).border = common_border
            sheet.cell(row=ending_row, column=2).border = common_border
            sheet[f"B{starting_row}"].alignment = px_styles.Alignment(horizontal="center", vertical="center")
            sheet.cell(row=ending_row + 1, column=2).border = spacing_border
            starting_row = ending_row + 2
            ending_row = starting_row + number_of_particles - 1
            iteration_number += 1
            # ! Checked the style of the last row
            # * Looks good 

        # * Appending the ARRAYS OF COORDINATES to the corresponding cell and styling them
        for row in range(optimization_df.shape[0]):
            for column in range(optimization_df.shape[1]):
                value = optimization_records[row][column]
                cell = sheet.cell(row=row + 4, column=column + 4)
                if isinstance(value, np.ndarray):
                    # * To avoid iterating through NaNs
                    value: str = ', '.join(map(str, value))
                    # * Makes each value of the tuple of floats a string (using map) and then joins them with a comma in a single string
                    cell.fill = px_styles.PatternFill(start_color=Color.cell2_fill, end_color=Color.cell2_fill, fill_type="solid") 
                    # ! Replace for a Color attribute once the branch is merged
                    # * Done
                    cell.border = common_border
                else:
                    cell.border = spacing_border
                cell.value = value
                cell.alignment = px_styles.Alignment(horizontal="center",
                    vertical="center", wrap_text=True)
                # * Rows and columns start at 1

        workbook.save(self.__xlsx_path)
        self.__particle_history.append(optimization_df)

    def create_spreadsheet(self) -> None:
        try:
            # * Initializing file as an Excel one
            empty_df: pd.DataFrame = pd.DataFrame()
            with pd.ExcelWriter(self.__xlsx_path, engine="openpyxl") as writer:
                empty_df.to_excel(excel_writer=writer)
                writer.book.active.title = "Optimization 1"
                writer.book.save(self.__xlsx_path)
        except FileExistsError as e:
            print(e)
            print("No need to create it again (should happen only in testing).")

    def print_optimization(self, optimization_index: int) -> None:
        pd.set_option("display.max_columns", None)
        pd.set_option("display.expand_frame_repr", False)
        pd.set_option("max_colwidth", None)
        pd.set_option("display.max_rows", None)
        print(self.__particle_history[optimization_index])
        
    def get_particle_history(self) -> list[pd.DataFrame]:
        return self.__particle_history